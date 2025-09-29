# This order converter app imports CSV files with Kit ID's and converts the order to a SKU based CSV upload.
# Created by Dusty Baker December 2023
# Updated by Dusty Baker May 2024 to add UPS shipping options, and EAGLE file manager.
#pyinstaller --onefile --noconsole --icon=assets/images/Eagle.ico --splash=assets/images/Splash.png Eagle.py
import os
import threading
from customtkinter import CTkProgressBar
from threading import Thread
from datetime import datetime
import tkinter as tk
from tkinter import *
from tkinter import filedialog, messagebox
from typing import Any
import customtkinter
import pandas as pd
import chardet
from PIL import Image
import pygame
import random
import webbrowser
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from pandas import Series, DataFrame
'''
import pyi_splash

pyi_splash.update_text("PyInstaller is a great software!")
pyi_splash.update_text("Second time's a charm!")

    # Close the splash screen. It does not matter when the call
    # to this function is made, the splash screen remains open until
    # this function is called or the Python program is terminated.
pyi_splash.close()
'''
customtkinter.set_appearance_mode("dark")  # Modes: system (default), light, dark
customtkinter.set_default_color_theme("dark-blue")  # Themes: blue (default), dark-blue, green

# Initialize pygame
pygame.mixer.init()

# Load the sound file
notification = pygame.mixer.Sound('assets/sounds/Propaganda.mp3')
intro = pygame.mixer.Sound('assets/sounds/LiberTea.mp3')
Error = pygame.mixer.Sound('assets/sounds/Error.mp3')

# Initialize error_added variable to keep track of whether an error message has been added
error_added = False

errors = []

# Function to check if files exist
def check_files_exist(file_paths):
    missing_files = [file for file in file_paths if not os.path.isfile(file)]
    if missing_files:
        messagebox.showerror("Missing Files", f"The following files are missing: {', '.join(missing_files)}")
        return False
    return True

# List of required files
required_files = [
    'assets/Templates/projects.csv',
    'assets/Templates/USCG_data.csv',
    'assets/Templates/USCG_WO_data.csv',
    'assets/images/Eagle.ico',

]

# Check if required files exist

if check_files_exist(required_files):
    # Try reading the CSV files with 'latin1' encoding first, then try 'ISO-8859-1' if it fails
    try:
        projects = pd.read_csv('assets/Templates/projects.csv', encoding='latin1')
        uscg_template = pd.read_csv('assets/Templates/USCG_data.csv', encoding='latin1')
        USCG_WO_template = pd.read_csv('assets/Templates/USCG_WO_data.csv', encoding='latin1')
    except UnicodeDecodeError:
        messagebox.showerror("Encoding Error", "Error reading the file with 'latin1' encoding. Trying 'ISO-8859-1'...")
        try:
            projects = pd.read_csv('assets/Templates/projects.csv', encoding='ISO-8859-1')
            uscg_template = pd.read_csv('assets/Templates/USCG_data.csv', encoding='ISO-8859-1')
            USCG_WO_template = pd.read_csv('assets/Templates/USCG_WO_data.csv', encoding='ISO-8859-1')
        except Exception as e:
            messagebox.showerror("File Error", f"An error occurred while reading the files: {e}")
else:
    # Exit the application if files are missing
    exit()
    
    
def read_file(file_path):
    if file_path.endswith('.csv'):
        # Detect encoding for CSV
        with open(file_path, 'rb') as f:
            result = chardet.detect(f.read())
        encoding = result['encoding']
        df = pd.read_csv(file_path, encoding=encoding)

    else:
        raise ValueError("Unsupported file format")
    return df

def save_as_utf8(df, output_path):
    df.to_csv(output_path, encoding='utf-8', index=False)


# Define the maximum number of characters in a column for shipment and WO inputs
allowable_lengths_for_shipments = {
    '2': 30,
    '4': 70,
    '5': 30,
    '6': 64,
    '7': 32,
    '8': 32,
    '9': 32,
    '10': 32,
    '11': 32,
    '12': 24,
    '13': 64,
    '14': 20,
    '16': 64,
    '17': 64,
    '18': 64,
    '19': 64,
    '20': 64,
    '21': 255,
    '23': 64,
    '24': 11,
    '25': 64,
    '26': 64,
    '27': 64,
    '28': 64,
    '29': 11,
    '30': 64,
    '31': 64,
    '32': 64,
    '33': 32,
    '34': 64,
    '35': 64,
}

# Define the maximum allowable lengths for ASN templates inputs
allowable_lengths_for_ASN = {
    '1': 20,
    '2': 64,
    '3': 12,
    '4': 20,
    '5': 64,
    '6': 64,
    '7': 20,
    '8': 64,
    '9': 64,
    '10': 64,
    '11': 64,
    '12': 64,
    '13': 64,
    '14': 64,
    '15': 64,
    '16': 64,
    '17': 64,

}



# Function to manipulate the input uscg data using the template
def manipulate_uscg_data_using_template(input_data, template):
    # Copy the first row under the header without manipulation
    manipulated_data = pd.DataFrame([input_data.iloc[0]])

    # Iterate over the rest of the input data starting from the second row
    for _, row in input_data.iloc[1:].iterrows():
        kit_id = row['14']
        matching_template_rows = template[template['Kit ID'] == kit_id]

        if matching_template_rows.empty:
            manipulated_data = pd.concat([manipulated_data, pd.DataFrame([row])])
        else:
            for _, template_row in matching_template_rows.iterrows():
                new_row = row.copy()

                # Update new_row with values from template_row where not NaN
                for column in template_row.index:
                    if column in row and pd.notna(template_row[column]):
                        new_row[column] = template_row[column]


                manipulated_data = pd.concat([manipulated_data, pd.DataFrame([new_row])])

    return manipulated_data


# Function to manipulate the input USCG data using the template for work orders
def manipulate_uscg_WO_data_using_template(input_data, template):
    # Copy the first row under the header without manipulation
    manipulated_data = pd.DataFrame([input_data.iloc[0]])

    # Iterate over the rest of the input data starting from the second row
    for _, row in input_data.iloc[1:].iterrows():
        # Use the correct column for Kit ID in input_data
        kit_id = row['Kit ID'] if 'Kit ID' in row else row['14']  # Use a fallback if needed

        # Ensure we're comparing using the correct column name in the template
        matching_template_rows = template[template['Kit ID'] == kit_id]

        if matching_template_rows.empty:
            manipulated_data = pd.concat([manipulated_data, pd.DataFrame([row])], ignore_index=True)
        else:
            for _, template_row in matching_template_rows.iterrows():
                new_row = row.copy()

                # Update new_row with values from template_row where not NaN
                for column in template_row.index:
                    if column in row and pd.notna(template_row[column]):
                        new_row[column] = template_row[column]

                manipulated_data = pd.concat([manipulated_data, pd.DataFrame([new_row])], ignore_index=True)

    # Group by column 12 and sum column 13
    manipulated_data['15'] = pd.to_numeric(manipulated_data['15'], errors='coerce')  # Ensure column 13 is numeric
    grouped_data = manipulated_data.groupby('14', as_index=False).agg({'15': 'sum'})

    # Merge the grouped data back to retain other columns, but only keep one row per unique '12' value
    result = manipulated_data.drop_duplicates(subset=['14']).drop(columns='15').merge(grouped_data, on='14')

    # Reorder columns to place column '13' immediately after column '12'
    cols = list(result.columns)
    col_12_index = cols.index('14')
    reordered_cols = cols[:col_12_index + 1] + ['15'] + cols[col_12_index + 1:-1]
    reordered_result: Series | None | DataFrame | Any = result[reordered_cols]

    return reordered_result


# Create a function that reads a CSV file into a dictionary and removes additional commas and line breaks
def check_and_remove_additional_commas(df: pd.DataFrame) -> pd.DataFrame:
    # Remove rows where all columns are empty
    df = df.dropna(how='all')

    # Define replacements, including newlines and carriage returns
    replacements = {
        ',': '',
        "'": '',
        '"': '',
        ';': '',
        ':': '',
        r'\(': '',
        r'\)': '',
        r'\[': '',
        r'\]': '',
        r'\{': '',
        r'\}': '',
        r'\?': '',
        r'\!': '',
        r'\n': '',   # Removes newline characters
        r'\r': ''    # Removes carriage return characters
    }

    # Apply the replacements only to string columns
    for col in df.select_dtypes(include=[object]).columns:
        df[col] = df[col].replace(replacements, regex=True)

    return df



# create a function that writes a cleaned CSV file
def write_cleaned_csv(output_file, cleaned_data):
    # Ensure cleaned_data is a DataFrame
    if not isinstance(cleaned_data, pd.DataFrame):
        cleaned_data = pd.DataFrame(cleaned_data)

    # Write the DataFrame to a CSV file
    cleaned_data.to_csv(output_file, index=False)


# Function to check the length of characters in a column
def check_character_length_shipment(df, length_dict, errors):
    for column, max_length in length_dict.items():
        if column in df.columns:
            # Check if any cell in the column exceeds the maximum allowable length
            exceeds_length = df[column].astype(str).apply(len) > max_length
            if exceeds_length.any():
                # Add an error with the index and value of the cell(s) that exceed the length
                error_rows = df[exceeds_length].index.tolist()
                error_values = df.loc[error_rows, column].tolist()
                for row, value in zip(error_rows, error_values):
                    errors.append(
                        f"Column '{column}' row {row + 2} exceeds allowable length of {max_length} characters."
                        f"\nError value: '{value}'")


# Function to check for the length of characters in a column for ASN's
def check_character_length_ASN(df, length_dict, errors):
    for column, max_length in length_dict.items():
        if column in df.columns:
            # Check if any cell in the column exceeds the maximum allowable length
            exceeds_length = df[column].astype(str).apply(len) > max_length
            if exceeds_length.any():
                # Add an error with the index and value of the cell(s) that exceed the length
                error_rows = df[exceeds_length].index.tolist()
                error_values = df.loc[error_rows, column].tolist()
                for row, value in zip(error_rows, error_values):
                    errors.append(
                        f"Column '{column}' row {row + 2} exceeds allowable length of {max_length} characters."
                        f"\nError value: '{value}'")


def USCG_Error_Handling(input_file):
    # Detect encoding for CSV
    with open(input_file, 'rb') as f:
        result = chardet.detect(f.read())
    encoding = result['encoding']


    # Read the CSV file into a DataFrame, skipping the first row after the header
    df = pd.read_csv(input_file, skiprows=[1], encoding=encoding)
    # Remove rows where all columns are empty
    df = df.dropna(how='all')



    # Define the valid kit IDs based on the template
    valid_kit_ids = uscg_template['Kit ID'].unique()

    # Initialize an empty list to store errors
    errors = []

    # Iterate over the rows in the DataFrame
    for index, row in df.iterrows():
        # Check if the row has more than 11 columns and if the 12th column value is not in valid_kit_ids
        if len(row) > 11 and row.iloc[13] not in valid_kit_ids:
            errors.append(
                f" row {index + 3}")  # Adding 3 to match the original line numbering

    return errors


# Function to convert USCG CSV using pandas and the template
def convert_USCG_csv(input_path, output_path, template):
    # Read the file (CSV or Excel) into the pandas DataFrame
    input_data = read_file(input_path)

    # Clean the commas from the input data
    cleaned_data = check_and_remove_additional_commas(input_data)
    write_cleaned_csv(input_path, cleaned_data)

    # Initialize errors list
    errors = []

    # Check the character length of the columns in the cleaned data
    check_character_length_shipment(input_data, allowable_lengths_for_shipments, errors)

    # If there are errors, raise a ValueError
    if errors:
        raise ValueError("\n".join(errors))

    # Manipulate the data using the template
    manipulated_data = manipulate_uscg_data_using_template(input_data, template)

    # Write the manipulated data to a new CSV file with UTF-8 encoding
    save_as_utf8(manipulated_data, output_path)


def sum_by_column_12(file_path):
    # Read the CSV file
    data = pd.read_csv(file_path)

    # Ensure column 15 contains numeric data
    data['15'] = pd.to_numeric(data['15'], errors='coerce')

    # Group by column 14 and sum the values in column 15
    result = data.groupby('14')['15'].sum().reset_index()

    # Return the result
    return result

# Function to convert USCG CSV using pandas and the template
def convert_USCG_WO_csv(input_path="Out.csv", output_path="USCG_WO_data.csv", template=USCG_WO_template):
    # Read the file (CSV or Excel) into the pandas DataFrame
    input_data = read_file(input_path)

    # Clean the commas from the input data
    cleaned_data = check_and_remove_additional_commas(input_data)
    write_cleaned_csv(input_path, cleaned_data)

    # Initialize errors list
    errors = []

    # Check the character length of the columns in the cleaned data
    check_character_length_shipment(input_data, allowable_lengths_for_shipments, errors)

    # If there are errors, raise a ValueError
    if errors:
        raise ValueError("\n".join(errors))

    # Manipulate the data using the template
    manipulated_data = manipulate_uscg_WO_data_using_template(input_data, template)

    # Write the manipulated data to a new CSV file with UTF-8 encoding
    save_as_utf8(manipulated_data, output_path)




# Function to convert the input file to a shipment
def USCG_convert_button_click():
    # Select the input file using a file dialog
    input_path = filedialog.askopenfilename(title="Select Input File for USCG", filetypes=[("CSV Files", "*.csv")])

    if not input_path:
        return  # User cancelled the file dialog

    # Check for correct USCG kit ID using the file path
    errors = USCG_Error_Handling(input_path)
    if errors:
        Error_window("ERROR -> Kit ID", f"There are incorrect Kit ID's in: {errors}")
        return

    try:
        # Read the file into a DataFrame after Kit ID validation
        input_data = read_file(input_path)

        # Clean the commas from the input data
        cleaned_data = check_and_remove_additional_commas(input_data)

        # Initialize errors list
        errors = []

        # Check the character length of the columns in the cleaned data
        check_character_length_shipment(cleaned_data, allowable_lengths_for_shipments, errors)

        # If there are errors, raise a ValueError
        if errors:
            raise ValueError("\n".join(errors))

    except (OSError, IOError, ValueError) as e:
        Error_window("Error", f"Error: {e}")
        return

    # Create the default output file path
    current_date = datetime.now().strftime("%m-%d-%Y")
    default_directory = r"T:\3PL Files\Shipment Template"
    default_filename = f"USCG Shipment {current_date}.csv"
    default_output_path = os.path.join(default_directory, default_filename)

    # Check if the default directory exists
    if not os.path.exists(default_directory):
        # Prompt the user to choose a file path if the default path is not available
        output_base_path = filedialog.asksaveasfilename(
            title="Save output file",
            defaultextension=".csv",
            filetypes=[("CSV Files", "*.csv")],
            initialfile=default_filename
        )
        if not output_base_path:
            return  # User cancelled the save dialog
    else:
        output_base_path = default_output_path

    try:
        # Call the conversion function
        convert_USCG_csv(input_path, output_base_path, uscg_template)
        Success_window(f"File saved successfully at\n{output_base_path}")
    except (OSError, IOError, ValueError) as e:
        Error_window("Error", f"Error: {e}")

# Function to convert the input file to a Work Order
def USCG_WO_button_click():
    # Select the input file using a file dialog
    input_path = filedialog.askopenfilename(title="Select Input File for USCG", filetypes=[("CSV Files", "*.csv")])

    if not input_path:
        return  # User cancelled the file dialog

    # Check for correct USCG kit ID using the file path
    errors = USCG_Error_Handling(input_path)
    if errors:
        Error_window("ERROR -> Kit ID", f"There are incorrect Kit ID's in: {errors}")
        return

    try:
        # Read the file into a DataFrame after Kit ID validation
        input_data = read_file(input_path)

        # Clean the commas from the input data
        cleaned_data = check_and_remove_additional_commas(input_data)

        # Initialize errors list
        errors = []

        # Check the character length of the columns in the cleaned data
        check_character_length_shipment(cleaned_data, allowable_lengths_for_shipments, errors)

        # If there are errors, stop execution
        if errors:
            Error_window("Character Length Error", f"Too many Characters, \n".join(errors))
            return

    except (OSError, IOError, ValueError) as e:
        Error_window("Error", f"Error: {e}")
        return

    # Create the default output file path
    current_date = datetime.now().strftime("%m-%d-%Y")
    default_directory = r"T:\3PL Files\Shipment Template"
    default_filename = f"USCG Work Order {current_date}.csv"
    default_output_path = os.path.join(default_directory, default_filename)

    # Check if the default directory exists
    if not os.path.exists(default_directory):
        # Prompt the user to choose a file path if the default path is not available
        output_base_path = filedialog.asksaveasfilename(
            title="Save output file",
            defaultextension=".csv",
            filetypes=[("CSV Files", "*.csv")],
            initialfile=default_filename
        )
        if not output_base_path:
            return  # User cancelled the save dialog
    else:
        output_base_path = default_output_path

    try:
        # Call the conversion function for Work Orders using USCG_WO_template instead of uscg_template
        convert_USCG_WO_csv(input_path, output_base_path, USCG_WO_template)


        Success_window(f"File saved successfully at\n{output_base_path}")

    except (OSError, IOError, ValueError) as e:
        Error_window("Error", f"Error: {e}")



# functinon to filter and save the input file as a shipment
def Eagle_shipment_button_click():
    input_path = filedialog.askopenfilename(title="Select Input File", filetypes=[("CSV Files", "*.csv")])

    if not input_path:
        return  # User cancelled the file dialog

    try:
        # Read the file into a DataFrame
        input_data = read_file(input_path)

        # Clean the commas from the input data
        cleaned_data = check_and_remove_additional_commas(input_data)
        write_cleaned_csv(input_path, cleaned_data)

        # Initialize errors list
        errors = []

        # Check the character length of the columns in the cleaned data
        check_character_length_shipment(cleaned_data, allowable_lengths_for_shipments, errors)

        # If there are errors, stop execution
        if errors:
            Error_window("Character Length Error", f"Too many Characters, \n".join(errors))
            return
    except (OSError, IOError, ValueError) as e:
        Error_window("Error", f"Error: {e}")
        return

    project_number = pd.read_csv(input_path, header=None, nrows=3).iloc[2, 0]
    project_number = str(project_number).strip()
    projects.columns = projects.columns.str.strip()
    projects['Project Number'] = projects['Project Number'].astype(str).str.strip()

    try:
        project_name_row = projects[projects['Project Number'] == project_number]
        if not project_name_row.empty:
            project_name = project_name_row['Project Name'].values[0]
        else:
            raise IndexError("Project number not found in DataFrame")
    except IndexError:
        Error_window("Project error", f"Project Number '{project_number}' not found.")
        return

    current_date = datetime.now().strftime("%m-%d-%Y")
    default_directory = r"T:\3PL Files\Shipment Template"
    default_filename = f"{project_name} Shipment {current_date}.csv"
    output_path = os.path.join(default_directory, default_filename)

    try:
        os.makedirs(default_directory, exist_ok=True)
        write_cleaned_csv(output_path, cleaned_data)
        Success_window(f"File saved successfully at \n{output_path}")
    except (OSError, IOError):
        output_base_path = filedialog.asksaveasfilename(
            title="Save Output File as Shipment",
            defaultextension=".csv",
            filetypes=[("CSV Files", "*.csv")],
            initialfile=default_filename
        )
        if not output_base_path:
            return
        write_cleaned_csv(output_base_path, cleaned_data)
        Success_window(f"File saved successfully at\n{output_base_path}")



# Function to convert the input file to a Work Order
def Eagle_WO_button_click():
    # Get the input file path
    input_path = filedialog.askopenfilename(title="Select Input File", filetypes=[("CSV Files", "*.csv")])

    if not input_path:
        return  # User cancelled the file dialog

    try:
        # Read the file into a DataFrame
        input_data = read_file(input_path)

        # Clean the commas from the input data
        cleaned_data = check_and_remove_additional_commas(input_data)
        write_cleaned_csv(input_path, cleaned_data)

        # Initialize errors list
        errors = []

        # Check the character length of the columns in the cleaned data
        check_character_length_shipment(cleaned_data, allowable_lengths_for_shipments, errors)

        # If there are errors, stop execution
        if errors:
            Error_window("Character Length Error", f"Too many Characters \n,".join(errors))
            return

        # Read the third row, first column to get the project number using pandas
        project_number = pd.read_csv(input_path, header=None, nrows=3).iloc[2, 0]

        # Convert project_number to string and strip any whitespace
        project_number = str(project_number).strip()

        # Ensure no extra spaces or unexpected characters in column names
        projects.columns = projects.columns.str.strip()

        # Convert the 'Project Number' column to string and strip any whitespace
        projects['Project Number'] = projects['Project Number'].astype(str).str.strip()

        # Map the project number to the project name using project dataframe from the projects.csv file using pandas
        try:
            project_name_row = projects[projects['Project Number'] == project_number]

            if not project_name_row.empty:
                project_name = project_name_row['Project Name'].values[0]

            else:
                raise IndexError("Project number not found in DataFrame")

        except IndexError:
            Error_window("Project Error", f"Project Number '{project_number}' not found.")
            return

        # Create the output file path
        current_date = datetime.now().strftime("%m-%d-%Y")
        default_directory = r"T:\3PL Files\Shipment Template"
        default_filename = f"{project_name} Work Order {current_date}.csv"
        output_path = os.path.join(default_directory, default_filename)

        # Save the cleaned CSV file with the appropriate name
        try:
            os.makedirs(default_directory, exist_ok=True)
            write_cleaned_csv(output_path, cleaned_data)
            Success_window(f"File saved successfully at\n{output_path}")

        except (OSError, IOError):
            # If the default path is unavailable, ask the user for the output directory and base filename
            output_base_path = filedialog.asksaveasfilename(
                title="Save Output File as Work Order",
                defaultextension=".csv",
                filetypes=[("CSV Files", "*.csv")],
                initialfile=default_filename
            )
            if not output_base_path:
                return  # User cancelled the save dialog

            # Save to the selected path
            write_cleaned_csv(output_base_path, cleaned_data)
            Success_window(f"File saved successfully at\n{output_base_path}")

    except (OSError, IOError, ValueError) as e:
        Error_window("Error", f"Error: {e}")



# Function to filter and save the input file as an ASN
def Eagle_ASN_button_click():
    # Get the input file path
    input_path = filedialog.askopenfilename(title="Select Input File", filetypes=[("CSV Files", "*.csv")])

    if not input_path:
        return  # User cancelled the file dialog

    try:
        # Read the file into a DataFrame
        input_data = read_file(input_path)

        # Clean the commas from the input data
        cleaned_data = check_and_remove_additional_commas(input_data)
        write_cleaned_csv(input_path, cleaned_data)

        # Initialize errors list
        errors = []

        # Check the character length of the columns in the cleaned data
        check_character_length_ASN(cleaned_data, allowable_lengths_for_shipments, errors)

        # If there are errors, stop execution
        if errors:
            Error_window("Character Length Error", f"Too many Characters, \n".join(errors))
            return

        # Read the fourth row, first column to get the project number using pandas
        project_number = pd.read_csv(input_path, header=None, nrows=4).iloc[3, 0]

        # Convert project_number to string and strip any whitespace
        project_number = str(project_number).strip()

        # Ensure no extra spaces or unexpected characters in column names
        projects.columns = projects.columns.str.strip()

        # Convert the 'Project Number' column to string and strip any whitespace
        projects['Project Number'] = projects['Project Number'].astype(str).str.strip()

        # Map the project number to the project name using project dataframe from the projects.csv file using pandas
        try:
            project_name_row = projects[projects['Project Number'] == project_number]

            if not project_name_row.empty:
                project_name = project_name_row['Project Name'].values[0]

            else:
                raise IndexError("Project number not found in DataFrame")

        except IndexError:
            Error_window("Project error", f"Project Number '{project_number}' not found.")
            return

        # Create the output file path
        current_date = datetime.now().strftime("%m-%d-%Y")
        default_directory = r"T:\3PL Files\ASN Template"
        default_filename = f"{project_name} ASN {current_date}.csv"
        output_path = os.path.join(default_directory, default_filename)

        # Save the cleaned CSV file with the appropriate name
        try:
            os.makedirs(default_directory, exist_ok=True)
            write_cleaned_csv(output_path, cleaned_data)
            Success_window(f"File saved successfully at\n{output_path}")

        except (OSError, IOError):
            # If the default path is unavailable, ask the user for the output directory and base filename
            output_base_path = filedialog.asksaveasfilename(
                title="Save Output File as ASN",
                defaultextension=".csv",
                filetypes=[("CSV Files", "*.csv")],
                initialfile=default_filename
            )
            if not output_base_path:
                return  # User cancelled the save dialog

            # Save to the selected path
            write_cleaned_csv(output_base_path, cleaned_data)
            Success_window(f"File saved successfully at\n{output_base_path}")

    except (OSError, IOError, ValueError) as e:
        Error_window("Error", f"Error: {e}")


# functinon to filter and save the input file as a shipment for B511
def B511_shipment_button_click():
    input_path = filedialog.askopenfilename(title="Select Input File", filetypes=[("CSV Files", "*.csv")])

    if not input_path:
        return  # User cancelled the file dialog

    try:
        # Read the file into a DataFrame
        input_data = read_file(input_path)

        # Clean the commas from the input data
        cleaned_data = check_and_remove_additional_commas(input_data)
        write_cleaned_csv(input_path, cleaned_data)


    except (OSError, IOError, ValueError) as e:
        Error_window("Error", f"Error: {e}")
        return

    project_number = pd.read_csv(input_path, header=None, nrows=4).iloc[3, 0]
    project_number = str(project_number).strip()
    projects.columns = projects.columns.str.strip()
    projects['Project Number'] = projects['Project Number'].astype(str).str.strip()

    try:
        project_name_row = projects[projects['Project Number'] == project_number]
        if not project_name_row.empty:
            project_name = project_name_row['Project Name'].values[0]
        else:
            raise IndexError("Project number not found in DataFrame")
    except IndexError:
        Error_window("Project error", f"Project Number '{project_number}' not found.")
        return

    current_date = datetime.now().strftime("%m-%d-%Y")
    default_directory = r"T:\InterfacesFiles\In"
    default_filename = f"{project_name} Shipment {current_date}.csv"
    output_path = os.path.join(default_directory, default_filename)

    try:
        os.makedirs(default_directory, exist_ok=True)
        write_cleaned_csv(output_path, cleaned_data)
        Success_window(f"File saved successfully at \n{output_path}")
    except (OSError, IOError):
        output_base_path = filedialog.asksaveasfilename(
            title="Save Output File as Shipment",
            defaultextension=".csv",
            filetypes=[("CSV Files", "*.csv")],
            initialfile=default_filename
        )
        if not output_base_path:
            return
        write_cleaned_csv(output_base_path, cleaned_data)
        Success_window(f"File saved successfully at\n{output_base_path}")

# Function to convert the input file to a Work Order
def B511_WO_button_click():
    # Get the input file path
    input_path = filedialog.askopenfilename(title="Select Input File", filetypes=[("CSV Files", "*.csv")])

    if not input_path:
        return  # User cancelled the file dialog

    try:
        # Read the file into a DataFrame
        input_data = read_file(input_path)

        # Clean the commas from the input data
        cleaned_data = check_and_remove_additional_commas(input_data)
        write_cleaned_csv(input_path, cleaned_data)


        # Read the third row, first column to get the project number using pandas
        project_number = pd.read_csv(input_path, header=None, nrows=4).iloc[3, 0]

        # Convert project_number to string and strip any whitespace
        project_number = str(project_number).strip()

        # Ensure no extra spaces or unexpected characters in column names
        projects.columns = projects.columns.str.strip()

        # Convert the 'Project Number' column to string and strip any whitespace
        projects['Project Number'] = projects['Project Number'].astype(str).str.strip()

        # Map the project number to the project name using project dataframe from the projects.csv file using pandas
        try:
            project_name_row = projects[projects['Project Number'] == project_number]

            if not project_name_row.empty:
                project_name = project_name_row['Project Name'].values[0]

            else:
                raise IndexError("Project number not found in DataFrame")

        except IndexError:
            Error_window("Project Error", f"Project Number '{project_number}' not found.")
            return

        # Create the output file path
        current_date = datetime.now().strftime("%m-%d-%Y")
        default_directory = r"T:\InterfacesFiles\In"
        default_filename = f"{project_name} Work Order {current_date}.csv"
        output_path = os.path.join(default_directory, default_filename)

        # Save the cleaned CSV file with the appropriate name
        try:
            os.makedirs(default_directory, exist_ok=True)
            write_cleaned_csv(output_path, cleaned_data)
            Success_window(f"File saved successfully at\n{output_path}")

        except (OSError, IOError):
            # If the default path is unavailable, ask the user for the output directory and base filename
            output_base_path = filedialog.asksaveasfilename(
                title="Save Output File as Work Order",
                defaultextension=".csv",
                filetypes=[("CSV Files", "*.csv")],
                initialfile=default_filename
            )
            if not output_base_path:
                return  # User cancelled the save dialog

            # Save to the selected path
            write_cleaned_csv(output_base_path, cleaned_data)
            Success_window(f"File saved successfully at\n{output_base_path}")

    except (OSError, IOError, ValueError) as e:
        Error_window("Error", f"Error: {e}")



# Function to filter and save the input file as an ASN
def B511_ASN_button_click():
    # Get the input file path
    input_path = filedialog.askopenfilename(title="Select Input File", filetypes=[("CSV Files", "*.csv")])

    if not input_path:
        return  # User cancelled the file dialog

    try:
        # Read the file into a DataFrame
        input_data = read_file(input_path)

        # Clean the commas from the input data
        cleaned_data = check_and_remove_additional_commas(input_data)
        write_cleaned_csv(input_path, cleaned_data)


        # Read the fourth row, first column to get the project number using pandas
        project_number = pd.read_csv(input_path, header=None, nrows=4).iloc[3, 0]

        # Convert project_number to string and strip any whitespace
        project_number = str(project_number).strip()

        # Ensure no extra spaces or unexpected characters in column names
        projects.columns = projects.columns.str.strip()

        # Convert the 'Project Number' column to string and strip any whitespace
        projects['Project Number'] = projects['Project Number'].astype(str).str.strip()

        # Map the project number to the project name using project dataframe from the projects.csv file using pandas
        try:
            project_name_row = projects[projects['Project Number'] == project_number]

            if not project_name_row.empty:
                project_name = project_name_row['Project Name'].values[0]

            else:
                raise IndexError("Project number not found in DataFrame")

        except IndexError:
            Error_window("Project error", f"Project Number '{project_number}' not found.")
            return

        # Create the output file path
        current_date = datetime.now().strftime("%m-%d-%Y")
        default_directory = r"T:\InterfacesFiles\In"
        default_filename = f"{project_name} ASN {current_date}.csv"
        output_path = os.path.join(default_directory, default_filename)

        # Save the cleaned CSV file with the appropriate name
        try:
            os.makedirs(default_directory, exist_ok=True)
            write_cleaned_csv(output_path, cleaned_data)
            Success_window(f"File saved successfully at\n{output_path}")

        except (OSError, IOError):
            # If the default path is unavailable, ask the user for the output directory and base filename
            output_base_path = filedialog.asksaveasfilename(
                title="Save Output File as ASN",
                defaultextension=".csv",
                filetypes=[("CSV Files", "*.csv")],
                initialfile=default_filename
            )
            if not output_base_path:
                return  # User cancelled the save dialog

            # Save to the selected path
            write_cleaned_csv(output_base_path, cleaned_data)
            Success_window(f"File saved successfully at\n{output_base_path}")

    except (OSError, IOError, ValueError) as e:
        Error_window("Error", f"Error: {e}")



        # Reporting


def autosize_columns(ws):
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column[0].column_letter].width = max_length + 2


def DFWR():
    # Local imports to ensure chart classes are available even if module-level changed
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.formatting.rule import DataBarRule

    input_path = filedialog.askopenfilename(
        title="Import Mantis Detail Inventory Report All Clients with Client Type",
        filetypes=[("Excel File", "*.xlsx")]
    )
    if not input_path:
        return

    # ---- Load data ----
    df = pd.read_excel(input_path, header=3)
    df.columns = df.columns.str.strip()

    # ---- Build Inventory Summary (grouped) for ws1 ----
    grouped_df = df.groupby(['Client', 'PO#', 'Order #', 'Description'], as_index=False)[
        ['Count Qty On Hand', 'Count Quantity Committed']
    ].sum()

    # Helper columns
    grouped_df.insert(1, 'Dell PM', '')
    grouped_df.insert(7, 'Shipping by EOM', '')
    grouped_df.insert(8, 'Full or Partial', '')

    # Merge project info
    grouped_df = grouped_df.merge(
        projects[['Project Name', 'Dell PM']],
        left_on='Client',
        right_on='Project Name',
        how='left'
    )
    grouped_df['Dell PM'] = grouped_df['Dell PM_y'].combine_first(grouped_df['Dell PM_x'])
    grouped_df = grouped_df.drop(columns=['Dell PM_x', 'Dell PM_y', 'Project Name'])

    # Reorder to put Dell PM after Client
    cols = grouped_df.columns.tolist()
    cols.insert(1, cols.pop(cols.index('Dell PM')))
    grouped_df = grouped_df[cols]

    # Numeric columns
    grouped_df['Count Quantity Committed'] = pd.to_numeric(grouped_df['Count Quantity Committed'], errors='coerce').fillna(0)
    grouped_df['Count Qty On Hand'] = pd.to_numeric(grouped_df['Count Qty On Hand'], errors='coerce').fillna(0)

    # Business logic
    grouped_df['Shipping by EOM'] = grouped_df['Count Quantity Committed'].apply(lambda x: 'N' if abs(x) < 1e-6 else 'Y')
    grouped_df['Full or Partial'] = grouped_df.apply(
        lambda row: 'TBD' if abs(row['Count Quantity Committed']) < 1e-6
        else ('Partial' if row['Count Quantity Committed'] < row['Count Qty On Hand'] else 'Full'),
        axis=1
    )

    # ---- Save intermediate file and load workbook ----
    output_file = "Dell Federal Weekly Reporting.xlsx"
    grouped_df.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    wb.worksheets[0].title = "Inventory Summary"
    ws1 = wb["Inventory Summary"]
    autosize_columns(ws1)

    # ---- Original data sheet (ws2) ----
    ws2 = wb.create_sheet(title="Mantis Detailed Inventory")
    for r in dataframe_to_rows(df, index=False, header=True):
        ws2.append(r)
    autosize_columns(ws2)

    # ---- Project Summary sheet (ws3) ----
    client_summary = grouped_df[['Client']].drop_duplicates().copy()
    client_summary = client_summary.merge(
        projects[['Project Name', 'Status']],
        left_on='Client',
        right_on='Project Name',
        how='left'
    ).drop(columns=['Project Name'])
    client_summary = client_summary[client_summary['Client'] != "Non Conforming"]

    ws3 = wb.create_sheet(title="Project Summary")
    for r in dataframe_to_rows(client_summary, index=False, header=True):
        ws3.append(r)
    autosize_columns(ws3)

    # ---------- Build the extended summary table on ws3 ----------
    # Ensure numeric source columns
    df['Count Qty On Hand'] = pd.to_numeric(df['Count Qty On Hand'], errors='coerce').fillna(0)

    # Column Q by position (17th column, index 16). If missing, use zeros to avoid errors.
    if df.shape[1] >= 17:
        col_q_series = pd.to_numeric(df.iloc[:, 16], errors='coerce').fillna(0)
    else:
        col_q_series = pd.Series([0] * len(df), index=df.index)
    df['__col_q__'] = col_q_series

    # Totals by client
    totals = (
        df.groupby('Client')['Count Qty On Hand']
        .sum()
        .reset_index()
        .rename(columns={'Count Qty On Hand': 'Inventory QTY'})
    )

    # Totals where column Q >= 90
    totals_q90 = (
        df[df['__col_q__'] >= 90]
        .groupby('Client')['Count Qty On Hand']
        .sum()
        .reset_index()
        .rename(columns={'Count Qty On Hand': 'Inventory QTY ≥90 Days'})
    )

    # Merge totals and clean
    inventory_summary = totals.merge(totals_q90, on='Client', how='left').fillna(0)
    inventory_summary = inventory_summary[
        (inventory_summary['Client'] != "Non Conforming") &
        (inventory_summary['Client'].notna()) &
        (inventory_summary['Client'].astype(str).str.strip() != "")
    ].sort_values(by='Inventory QTY', ascending=False).reset_index(drop=True)

    # % column
    inventory_summary['% ≥90'] = (
        inventory_summary['Inventory QTY ≥90 Days'] / inventory_summary['Inventory QTY']
    ).replace([float('inf'), float('-inf')], 0).fillna(0)

    # ---- Write the table headers ----
    start_row = ws3.max_row + 3
    ws3.cell(row=start_row, column=1, value="Client")
    ws3.cell(row=start_row, column=2, value="Inventory QTY")
    ws3.cell(row=start_row, column=3, value="Inventory QTY ≥90 Days")
    ws3.cell(row=start_row, column=4, value="% ≥90")

    # Header styling for this table
    header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    for col in range(1, 5):
        c = ws3.cell(row=start_row, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center_align

    # ---- Write data rows ----
    for i, row in inventory_summary.iterrows():
        ws3.cell(row=start_row + i + 1, column=1, value=row['Client'])
        ws3.cell(row=start_row + i + 1, column=2, value=row['Inventory QTY'])
        ws3.cell(row=start_row + i + 1, column=3, value=row['Inventory QTY ≥90 Days'])
        pct_cell = ws3.cell(row=start_row + i + 1, column=4, value=float(row['% ≥90']))
        pct_cell.number_format = '0.0%'

    # Ranges for charts
    data_start = start_row + 1
    data_end = start_row + len(inventory_summary)

    # Optional: data bars on % ≥90 to quickly scan
    if data_end >= data_start:
        db_rule = DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="63BE7B")
        ws3.conditional_formatting.add(f"D{data_start}:D{data_end}", db_rule)

    # ---- Charts (only if there’s data) ----
    if data_end >= data_start:
        # --- Pie (based on total inventory) ---
        pie = PieChart()
        pie.title = "Inventory Distribution by Project"
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showVal = False
        pie.dataLabels.showCatName = False
       
        pie.dataLabels.showLeaderLines = True  # <-- add leader lines

        pie_data = Reference(ws3, min_col=2, min_row=data_start, max_row=data_end)   # Inventory QTY
        pie_labels = Reference(ws3, min_col=1, min_row=data_start, max_row=data_end) # Client
        pie.add_data(pie_data, titles_from_data=False)
        pie.set_categories(pie_labels)

        # Place the pie chart
        ws3.add_chart(pie, "H2")
        # --- Simple Bar: Total Inventory by Client (absolute) ---
        abs_bar = BarChart()
        abs_bar.type = "col"
        abs_bar.grouping = "clustered"
        abs_bar.title = "Total Inventory by Project"
        abs_bar.y_axis.title = "Quantity"
        abs_bar.x_axis.title = "Client"

        # Series = ONLY Inventory QTY (col 2); include header row for series title
        abs_data = Reference(ws3, min_col=2, max_col=2, min_row=start_row, max_row=data_end)
        abs_cats = Reference(ws3, min_col=1, min_row=data_start, max_row=data_end)

        abs_bar.add_data(abs_data, titles_from_data=True)
        abs_bar.set_categories(abs_cats)

        # Show values on bars
        abs_bar.dataLabels = DataLabelList()
        abs_bar.dataLabels.showVal = True

        # Optional cosmetics
        abs_bar.legend = None  # single series doesn't need a legend
        abs_bar.varyColors = False

        # Place chart so it doesn't overlap others
        ws3.add_chart(abs_bar, "H16" )

    # ---- Reorder tabs ----
    desired_order = ["Project Summary", "Inventory Summary", "Mantis Detailed Inventory"]
    wb._sheets = [wb[s] for s in desired_order]

    # ---- Global formatting (center + borders + header row 1) ----
    center_align_all = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = center_align_all
                    cell.border = thin_border
        for cell in sheet[1]:
            if cell.value is not None:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align_all
                cell.border = thin_border
        sheet.freeze_panes = "A2"

    # ---- Save dialog ----
    today_str = datetime.today().strftime("%B %d, %Y")
    default_filename = f"Dell Federal Weekly Reporting ({today_str}).xlsx"
    output_path = filedialog.asksaveasfilename(
        title="Save Report As",
        defaultextension=".xlsx",
        initialfile=default_filename,
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not output_path:
        return

    wb.save(output_path)
    Success_window("File saved successfully")

        






# Open the assets folder and allow the user to open a .csv file from "assets/Templates"
def modify_template():
    Caution_window("CAUTION", "Please make sure to save the file in the same format, name and location as the original template.\n"
                              "Any changes made to the file may affect the functionality of the application.")
    file_path = filedialog.askopenfilename(title="Select a Template File", initialdir="P:\Eagle_assets/Templates",
                                           filetypes=[("CSV Files", "*.csv")])
    if file_path:
        try:
            os.system(f'start excel "{file_path}"')
        except (OSError, IOError) as e:
            Error_window("Error", f"Error: {e}")




# Open file dialog to allow user to select a file to add as a meme
def Add_meme():
    file_path = filedialog.askopenfilename(title="Select a Dank Meme you would like to add",
                                           filetypes=[("Image Files", "*.png;*.jpg;*.jpeg;*.gif")])
    # save the image to the assets/memes folder
    if file_path:
        file_name = os.path.basename(file_path)
        output_path = os.path.join("assets/memes", file_name)
        try:
            os.makedirs("assets/memes", exist_ok=True)
            shutil.copy(file_path, output_path)
            Success_window(f"File saved successfully at\n{output_path}")
        except (OSError, IOError) as e:
            Error_window("Error", f"Error: {e}")




def theme_selection(theme):
    if theme == 0:
        customtkinter.set_appearance_mode("light")
        update_menu_colors()
    else :
        customtkinter.set_appearance_mode("dark")
        update_menu_colors()


def help_us():
    # Open the email client with new email
    email_address = "dbaker@premierss.com"
    subject = "Eagle File Manager Help Request"
    body = "The user has requested help with the Eagle File Manager. Please provide assistance."
    webbrowser.open(f"mailto:{email_address}?subject={subject}&body={body}")





    pass


def update_menu_colors():
    appearance_mode = customtkinter.get_appearance_mode()
    if appearance_mode == "Dark":
        fg_color = "#333333"
        fg_text = "#ffffff"
    else:
        fg_color = "#f0f0f0"
        fg_text = "#000000"

    menu_frame.configure(fg_color=fg_color)
    for widget in menu_frame.winfo_children():
        if isinstance(widget, customtkinter.CTkButton):
            widget.configure(fg_color=fg_color, text_color=fg_text)

def show_submenu(menu, button):
    menu.tk_popup(button.winfo_rootx(), button.winfo_rooty() + button.winfo_height())

# Function to center the window
def center_window(window, width, height):
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x = int((screen_width / 2) - (width / 2))
    y = int((screen_height / 2) - (height / 2))
    window.geometry(f'{width}x{height}+{x}+{y}')



# GUI_______________________________________________________________________________________________________

root = customtkinter.CTk()
root.title("Eagle File Manager 1.9 -Dell Federal Weekly Reporting\u00AE added!")
root.geometry("600x650")
root.iconbitmap('assets/images/Eagle.ico')
#locate the main window in the center of the screen
window_width = 600
window_height = 650
center_window(root, window_width, window_height)

# Assuming image_path is defined
image_path = "assets/images/Eagle.ico"


# Create a custom frame to act as the menu bar that is only tall enough to hold the buttons
menu_frame = customtkinter.CTkFrame(root,height=10, fg_color="transparent")
menu_frame.pack(side="top", fill="x",)

def create_menu_button(text, menu):
    button = customtkinter.CTkButton(menu_frame, text=text, command=lambda: show_submenu(menu, button), corner_radius=0, fg_color="transparent", text_color="#FFFFFF")
    button.pack(side="left")
    return button

file_menu = tk.Menu(menu_frame, tearoff=0)
file_menu.add_command(label="Modify Template", command=modify_template)
file_menu.add_command(label="Add Meme", command=Add_meme)
file_menu.add_separator()
file_menu.add_command(label="Exit", command=root.quit)


setting_menu = tk.Menu(menu_frame, tearoff=0)
setting_menu.add_command(label="Light theme", command=lambda: theme_selection(0))
setting_menu.add_command(label="Dark theme", command=lambda: theme_selection(1))


file_button = create_menu_button("File", file_menu)
setting_button = create_menu_button("Setting", setting_menu)
help_button = customtkinter.CTkButton(menu_frame, text="Help", command=help_us, corner_radius=0, fg_color="transparent", text_color="#FFFFFF")
help_button.pack(side="left")


# create a tab view with custom tkinter that fills root window
My_tab = customtkinter.CTkTabview(root)
My_tab.pack(fill="both", expand=True, side="top")

# create a tab
tab_1 = My_tab.add("The Eagle")
tab_2 = My_tab.add("USCG")
tab_3 = My_tab.add("B511")
tab_4 = My_tab.add("Reporting")


# TAB 1 / The EAGLE_________________________________________________________________________________________
## Background image for tab_1 / Eagle
Eagle_image = customtkinter.CTkImage(light_image=Image.open('assets/images/Golden_Eagle.png'),
                                     dark_image=Image.open('assets/images/Golden_Eagle.png'),
                                     size=(400, 400))

Eagle_label = customtkinter.CTkLabel(tab_1, text="", image=Eagle_image)
Eagle_label.pack(side='top', pady=10)

# Create a button that calls the Eagle_WO_button_click function for Eagle csv filtering and save as a Work Order
Work_order_button_Eagle = customtkinter.CTkButton(tab_1,
                                                  text="Filter and save Work Order", command=Eagle_WO_button_click,
                                                  border_width=2, border_color="#a73a3a", fg_color="#1a1b1d",hover_color="#764f33")
Work_order_button_Eagle.pack(side='bottom', pady=10)

# Create a button that calls the Eagle_shipment_button_click function for Eagle csv filtering and save as shipment
Shipment_button_Eagle = customtkinter.CTkButton(tab_1,
                                                text="Filter and save Shipment",
                                                command=Eagle_shipment_button_click,
                                                border_width=2, border_color="#a73a3a", fg_color="#1a1b1d",hover_color="#764f33")
Shipment_button_Eagle.pack(side='bottom', pady=10)

# Create a button that called the Eagle_ASN_button_click function for Eagle csv filtering and save as ASN
ASN_button_Eagle = customtkinter.CTkButton(tab_1,
                                           text="Filter and save ASN", command=Eagle_ASN_button_click,
                                           border_width=2, border_color="#a73a3a", fg_color="#1a1b1d",hover_color="#764f33")
ASN_button_Eagle.pack(side='bottom', pady=10)

# TAB 2 / USCG_____________________________________________________________________________________________
# background image for tab_2 / USCG
USCG_image = customtkinter.CTkImage(light_image=Image.open('assets/images/background.png'),
                                    dark_image=Image.open('assets/images/background.png'),
                                    size=(300, 300))

USCG_label = customtkinter.CTkLabel(tab_2, text="", image=USCG_image)
USCG_label.pack(side='top', pady=20)

# Create a button that calls the convert_shipment_csv function for USCG
convert_button = customtkinter.CTkButton(tab_2,
                                         text="Convert Shipment", command=USCG_convert_button_click,
                                         border_width=2, border_color="gold")
convert_button.pack(side='bottom', pady=20)

#Create a button that calls the convert_Work_Order_csv function for USCG
convert_button = customtkinter.CTkButton(tab_2,
                                         text="Convert Work Order", command=USCG_WO_button_click,
                                         border_width=2, border_color="gold")
convert_button.pack(side='bottom', pady=20)


# TAB 3 / B511_____________________________________________________________________________________________
# Background image for tab_3 / B511
B511_image = customtkinter.CTkImage(light_image=Image.open('assets/images/eagle.jpg'),
                                    dark_image=Image.open('assets/images/eagle.jpg'),
                                    size=(400, 400))
# Create a label for the B511 image
B511_label = customtkinter.CTkLabel(tab_3, text="", image=B511_image)
B511_label.pack(side='top', pady=10)

# Create a button that calls the convert_Work_Order_csv function for B511
convert_button = customtkinter.CTkButton(tab_3,
                                            text="Filter and save Work Order", command=B511_WO_button_click,
                                            border_width=2, border_color="#a73a3a", fg_color="#1a1b1d",hover_color="#764f33")
convert_button.pack(side='bottom', pady=10)


# Create a button that calls the convert_shipment_csv function for B511
convert_button = customtkinter.CTkButton(tab_3,
                                            text="Filter and save Shipment", command=B511_shipment_button_click,
                                            border_width=2, border_color="#a73a3a", fg_color="#1a1b1d",hover_color="#764f33")
convert_button.pack(side='bottom', pady=10)



# Create a button that calls the convert_ASN_csv function for B511
convert_button = customtkinter.CTkButton(tab_3,
                                            text="Filter and save ASN", command=B511_ASN_button_click,
                                            border_width=2, border_color="#a73a3a", fg_color="#1a1b1d",hover_color="#764f33")
convert_button.pack(side='bottom', pady=10)


# TAB 4 / Reporting_____________________________________________________________________________________________________
# Create a button that calls the Dell Federal Weekly reporting function

def run_dfwr_with_loading():
    progress_bar = CTkProgressBar(master=root)
    progress_bar.pack(pady=10)
    progress_bar.start()

    def task():
        try:
            DFWR()
        finally:
            # Stop progress bar on completion
            progress_bar.stop()
            progress_bar.destroy()

    Thread(target=task).start()

DFWR_Button = customtkinter.CTkButton(tab_4,
                                            text="Dell Federal Weekly Reporting", command=run_dfwr_with_loading,
                                            border_width=2, border_color="#a73a3a", fg_color="#1a1b1d",hover_color="#764f33")
DFWR_Button.pack(side='bottom', pady=10)



# custom Tkinter top level window for success message
def Success_window(message):
    s_window = customtkinter.CTkToplevel()
    s_window.title("An Eagle never misses")
    s_window.geometry("500x200")
    s_window.attributes('-topmost', True)
    s_window.iconbitmap('assets/images/Eagle.ico')

    #locate the window in the center of the screen
    window_width = 500
    window_height = 200
    center_window(s_window, window_width, window_height)

    # Show the eagle image
    Eagle_image = customtkinter.CTkImage(light_image=Image.open('assets/images/eagle.ico'),
                                         dark_image=Image.open('assets/images/eagle.ico'),
                                         size=(100, 100))

    # Create a frame to hold the image and text
    content_frame = customtkinter.CTkFrame(s_window)
    content_frame.pack(expand=True, fill='both', padx=10, pady=10)

    # Place the eagle image on the left side
    Eagle_label = customtkinter.CTkLabel(content_frame, text="", image=Eagle_image)
    Eagle_label.pack(side='left', padx=10, pady=10)

    # Create a label for the success message, placed to the right of the image
    success_label = customtkinter.CTkLabel(content_frame, text=message, wraplength=300)
    success_label.pack(side='left', padx=10, pady=10, expand=True, fill='both')

    # Create a button centered at the bottom
    Create_button = customtkinter.CTkButton(s_window, text="Lets GO!", command=s_window.destroy)
    Create_button.pack(side='bottom', pady=20)

    # Play the notification sound
    notification.play()


#custom Tkinter top level window for error message
# Custom Tkinter top level window for error message
def Error_window(title, message):
    e_window = customtkinter.CTkToplevel(root)
    e_window.title(title)
    e_window.geometry("600x300")  # Increase height to make room for the button
    e_window.iconbitmap('assets/images/Eagle.ico')  # Set the icon for the error window
    e_window.attributes('-topmost', True)
    #locate the window in the center of the screen
    window_width = 600
    window_height = 300
    center_window(e_window, window_width, window_height)

    # Select a random image from the "assets/memes" folder
    memes_folder = 'assets/memes'
    meme_files = [f for f in os.listdir(memes_folder) if os.path.isfile(os.path.join(memes_folder, f))]
    random_meme = os.path.join(memes_folder, random.choice(meme_files))

    # Show the random meme image
    Meme_image = customtkinter.CTkImage(light_image=Image.open(random_meme),
                                        dark_image=Image.open(random_meme),
                                        size=(200, 200))

    # Create a frame to hold the image and text
    content_frame = customtkinter.CTkFrame(e_window)
    content_frame.pack(expand=True, fill='both', padx=10, pady=10)

    # Place the meme image on the left side
    Meme_label = customtkinter.CTkLabel(content_frame, text="", image=Meme_image)
    Meme_label.pack(side='left', padx=10, pady=10)

    # Create a scrollable frame for the error message
    scrollable_frame = customtkinter.CTkScrollableFrame(content_frame, width=300, height=100)  # Reduce height
    scrollable_frame.pack(side='left', padx=10, pady=10, expand=True, fill='both')

    # Create a label for the error message inside the scrollable frame
    error_label = customtkinter.CTkLabel(scrollable_frame, text=message, wraplength=280)
    error_label.pack(padx=10, pady=10, expand=True, fill='both')

    # Create a button centered at the bottom
    Create_button = customtkinter.CTkButton(e_window, text="Disappointment", command=e_window.destroy)
    Create_button.pack(side='bottom', pady=10)  # Adjust pady to make room

    # Play the error sound
    Error.play()

def Caution_window(title, message):
    c_window = customtkinter.CTkToplevel()
    c_window.title(title)
    c_window.geometry("500x200")
    c_window.attributes('-topmost', True)
    c_window.iconbitmap('assets/images/Eagle.ico')
    window_width = 500
    window_height = 200
    center_window(c_window, window_width, window_height)

    # Show the eagle image
    Eagle_image = customtkinter.CTkImage(light_image=Image.open('assets/images/eagle.ico'),
                                         dark_image=Image.open('assets/images/eagle.ico'),
                                         size=(100, 100))

    # Create a frame to hold the image and text
    content_frame = customtkinter.CTkFrame(c_window)
    content_frame.pack(expand=True, fill='both', padx=10, pady=10)

    # Place the eagle image on the left side
    Eagle_label = customtkinter.CTkLabel(content_frame, text="", image=Eagle_image)
    Eagle_label.pack(side='left', padx=10, pady=10)

    # Create a label for the success message, placed to the right of the image
    success_label = customtkinter.CTkLabel(content_frame, text=message, wraplength=300)
    success_label.pack(side='left', padx=10, pady=10, expand=True, fill='both')

    # Create a button centered at the bottom
    Create_button = customtkinter.CTkButton(c_window, text="I 100% understand", command=c_window.destroy)
    Create_button.pack(side='bottom', pady=20)

    # Play the notification sound
    notification.play()


def main():
    intro.play()
    #set timer for the intro sound to fade out
    intro.fadeout(15000)
    root.mainloop()






if __name__ == "__main__":
    main()
